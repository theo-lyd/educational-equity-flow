"""XLSX ingestion to normalized Bronze records."""

from __future__ import annotations

import warnings
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from .normalizers import infer_quality, normalize_ags, normalize_label, parse_numeric


def ingest_xlsx(path: Path) -> pl.DataFrame:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v).strip() for v in row])

    wb.close()

    data_start = 0
    for i, row in enumerate(rows):
        joined = " ".join(row)
        if any(x in joined for x in ["Deutschland", "Geisteswissenschaften", "Vorschulbereich"]):
            data_start = i
            break

    header_rows = rows[max(data_start - 3, 0) : data_start]
    width = max((len(r) for r in rows[data_start:]), default=0)

    metric_names: list[str] = []
    for col in range(width):
        parts: list[str] = []
        for hr in header_rows:
            if col < len(hr):
                v = normalize_label(hr[col])
                if v and v not in {"Anzahl", "WS 2023/24"}:
                    parts.append(v)
        metric_names.append(" | ".join(dict.fromkeys(parts)) if parts else f"col_{col}")

    records: list[dict[str, object]] = []
    current_year: str | None = None
    current_ags: str | None = None
    current_region: str | None = None

    for row in rows[data_start:]:
        row = row + [""] * (width - len(row))

        if row[0].isdigit() and len(row[0]) == 4:
            current_year = row[0]
            continue

        if row[0] and row[0] not in {"DG", "DL"}:
            current_ags = normalize_ags(row[0])
        if row[1]:
            current_region = normalize_label(row[1])

        for idx in range(3, width):
            raw = row[idx]
            value = parse_numeric(raw)
            records.append(
                {
                    "dataset": path.stem,
                    "source_file": path.name,
                    "year": current_year,
                    "ags": current_ags,
                    "region": current_region,
                    "dimension_1": normalize_label(row[2]) if len(row) > 2 else None,
                    "dimension_2": None,
                    "dimension_3": None,
                    "metric_name": metric_names[idx],
                    "raw_value": raw,
                    "value": value,
                    "quality": infer_quality(raw),
                }
            )

    return pl.DataFrame(records)
