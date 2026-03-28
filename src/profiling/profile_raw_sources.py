"""Generate Phase 03 raw-data profiling artifacts and source contracts."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


def guess_encoding(path: Path) -> str:
    blob = path.read_bytes()[:200_000]
    for enc in ("utf-8-sig", "utf-8", "iso-8859-1", "latin-1"):
        try:
            blob.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


def read_text_lines(path: Path, encoding: str, max_lines: int = 200) -> list[str]:
    lines: list[str] = []
    with path.open("r", encoding=encoding, errors="replace") as fh:
        for _, line in zip(range(max_lines), fh):
            lines.append(line.rstrip("\n"))
    return lines


def detect_delimiter(lines: list[str]) -> str:
    candidates = [";", "\t", ",", "|"]
    score = {c: 0 for c in candidates}
    for line in lines:
        for c in candidates:
            score[c] += line.count(c)
    return max(score, key=score.get)


def normalize_colname(value: str) -> str:
    v = value.strip()
    v = re.sub(r"\s+", "_", v)
    return re.sub(r"[^A-Za-z0-9_\-]", "", v)


def detect_header_index(lines: list[str], delimiter: str) -> int:
    for i, line in enumerate(lines):
        low = line.lower()
        if "ags" in low and ("jahr" in low or delimiter in line):
            return i
    for i, line in enumerate(lines):
        if line.count(delimiter) >= 4:
            return i
    return 0


def detect_data_start(lines: list[str], header_index: int, delimiter: str) -> int:
    year_re = re.compile(r"\b(19|20)\d{2}\b")
    ags_re = re.compile(r"\b\d{5}\b")
    for i in range(header_index + 1, len(lines)):
        line = lines[i]
        if not line.strip():
            continue
        if year_re.search(line) or ags_re.search(line):
            return i
    return max(header_index + 1, 0)


def infer_value_type(samples: list[str]) -> str:
    non_empty = [s for s in samples if s not in ("", "-", ".", "x", "X")]
    if not non_empty:
        return "empty_or_marker"

    if all(re.fullmatch(r"\d{4}", s) for s in non_empty[:100]):
        return "year"
    if all(re.fullmatch(r"\d+", s) for s in non_empty[:100]):
        return "integer"
    if all(re.fullmatch(r"\d+[\.,]\d+", s) for s in non_empty[:100]):
        return "decimal"
    if all(re.fullmatch(r"\d{5,8}", s) for s in non_empty[:100]):
        return "code"
    return "string"


def profile_csv(path: Path) -> dict[str, Any]:
    encoding = guess_encoding(path)
    lines = read_text_lines(path, encoding=encoding)
    delimiter = detect_delimiter(lines)
    header_idx = detect_header_index(lines, delimiter)
    data_start_idx = detect_data_start(lines, header_idx, delimiter)

    header_parts = [normalize_colname(x) for x in lines[header_idx].split(delimiter)]
    header_parts = [h for h in header_parts if h]

    sample_rows: list[list[str]] = []
    for line in lines[data_start_idx : data_start_idx + 60]:
        sample_rows.append([p.strip() for p in line.split(delimiter)])

    col_samples: dict[str, list[str]] = {}
    for row in sample_rows:
        for i, value in enumerate(row):
            if i >= len(header_parts):
                continue
            col_samples.setdefault(header_parts[i], []).append(value)

    inferred_types = {k: infer_value_type(v) for k, v in col_samples.items()}

    quality_markers = Counter()
    for row in sample_rows:
        for value in row:
            if value.strip() in {"-", ".", "x", "X", "e"}:
                quality_markers[value.strip()] += 1

    metadata_lines = max(data_start_idx - header_idx - 1, 0)

    return {
        "file_type": "csv",
        "encoding": encoding,
        "delimiter": delimiter,
        "header_row_index": header_idx,
        "data_start_row_index": data_start_idx,
        "metadata_lines_between_header_and_data": metadata_lines,
        "columns": header_parts,
        "inferred_types": inferred_types,
        "quality_markers_seen": dict(quality_markers),
    }


def profile_xlsx(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 120:
            break
        rows.append(["" if v is None else str(v).strip() for v in row])

    header_idx = 0
    for i, row in enumerate(rows):
        low = " ".join(row).lower()
        if "ags" in low and "jahr" in low:
            header_idx = i
            break
        if sum(1 for c in row if c) >= 5:
            header_idx = i
            break

    columns = [normalize_colname(c) for c in rows[header_idx] if c]
    data_preview = rows[header_idx + 1 : header_idx + 8]

    wb.close()

    return {
        "file_type": "xlsx",
        "sheet_name": ws.title,
        "header_row_index": header_idx,
        "columns": columns,
        "sample_data_rows": data_preview,
    }


def profile_xml(path: Path) -> dict[str, Any]:
    tag_counter: Counter[str] = Counter()
    axes: set[str] = set()
    variables: set[str] = set()
    quality_markers: Counter[str] = Counter()
    coordinate_samples: list[str] = []

    for event, elem in ET.iterparse(path, events=("start",)):
        tag = elem.tag.split("}")[-1]
        tag_counter[tag] += 1

        if tag == "AXIS":
            var_name = elem.attrib.get("VARIABLE", "")
            if var_name:
                axes.add(var_name)
        if tag == "VARIABLE":
            var_name = elem.attrib.get("NAME", "")
            if var_name:
                variables.add(var_name)
        if tag == "VALUE":
            q = elem.attrib.get("QUALITY", "")
            if q:
                quality_markers[q] += 1
            if len(coordinate_samples) < 8 and "COORDINATE" in elem.attrib:
                coordinate_samples.append(elem.attrib["COORDINATE"])

    return {
        "file_type": "xml",
        "top_tags": dict(tag_counter.most_common(12)),
        "axes_detected": sorted(axes),
        "variables_detected": sorted(variables),
        "quality_markers_seen": dict(quality_markers),
        "coordinate_samples": coordinate_samples,
    }


def build_contract(file_name: str, profile: dict[str, Any]) -> dict[str, Any]:
    lower_cols = [c.lower() for c in profile.get("columns", [])]
    dimension_candidates = [
        c
        for c in ["ags", "kreise", "jahr", "gesins", "hs-fg2", "deut_ausl", "geschlecht"]
        if any(c in col for col in lower_cols)
    ]

    contract = {
        "file": file_name,
        "file_type": profile["file_type"],
        "required_dimension_candidates": sorted(set(dimension_candidates)),
        "required_rules": {
            "must_include_geo_key": ["AGS", "KREISE"],
            "must_include_time_key": ["JAHR"],
            "accepted_quality_markers": ["-", ".", "x", "X", "e"],
        },
        "ingestion_hints": {
            "normalize_ags_to_5_digits": True,
            "normalize_abbreviations": ["dar.", "Mio", "K"],
            "drop_metadata_rows_before_header": True,
        },
    }

    if profile["file_type"] == "xml":
        contract["required_dimension_candidates"] = profile.get("axes_detected", [])

    return contract


def run(raw_dir: Path, out_dir: Path) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(
        [p for p in raw_dir.iterdir() if p.is_file() and p.name != "RAW_STATE_LOCK.csv"],
        key=lambda p: p.name,
    )

    profiles: dict[str, dict[str, Any]] = {}
    contracts: dict[str, dict[str, Any]] = {}

    for file_path in files:
        if file_path.suffix.lower() == ".csv":
            profile = profile_csv(file_path)
        elif file_path.suffix.lower() == ".xlsx":
            profile = profile_xlsx(file_path)
        elif file_path.suffix.lower() == ".xml":
            profile = profile_xml(file_path)
        else:
            continue

        profiles[file_path.name] = profile
        contracts[file_path.name] = build_contract(file_path.name, profile)

    summary = {
        "generated_utc": datetime.now(UTC).isoformat(),
        "raw_directory": str(raw_dir),
        "profile_count": len(profiles),
        "profiles": profiles,
        "contracts": contracts,
    }

    (out_dir / "schema_snapshots.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8"
    )
    (out_dir / "source_contracts.json").write_text(
        json.dumps(contracts, indent=2, ensure_ascii=True), encoding="utf-8"
    )

    lines = [
        "# Phase 03 Raw Profiling Summary",
        "",
        f"Generated (UTC): {summary['generated_utc']}",
        "",
        "## File Summaries",
        "",
    ]

    for name, p in profiles.items():
        lines.append(f"### {name}")
        lines.append(f"- Type: {p['file_type']}")
        if "encoding" in p:
            lines.append(f"- Encoding: {p['encoding']}")
        if "delimiter" in p:
            lines.append(f"- Delimiter: `{p['delimiter']}`")
        if "header_row_index" in p:
            lines.append(f"- Header row index: {p['header_row_index']}")
        if "data_start_row_index" in p:
            lines.append(f"- Data start row index: {p['data_start_row_index']}")
        if "columns" in p:
            preview = ", ".join(p["columns"][:12])
            lines.append(f"- Column preview: {preview}")
        if p["file_type"] == "xml":
            lines.append(f"- Axes detected: {', '.join(p.get('axes_detected', []))}")
        lines.append("")

    (out_dir / "RAW_PROFILING_SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")

    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate Phase 03 profiling artifacts")
    parser.add_argument("--raw-dir", default="data/raw")
    parser.add_argument("--out-dir", default="docs/phase_03/artifacts")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    summary = run(raw_dir=Path(args.raw_dir), out_dir=Path(args.out_dir))
    print(
        "Phase 03 profiling complete:",
        f"profiles={summary['profile_count']}",
        f"out_dir={args.out_dir}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
